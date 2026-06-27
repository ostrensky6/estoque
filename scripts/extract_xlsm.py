"""Extrai Laboratorio1.xlsm -> seed/seed.sql + CSVs.
Normaliza codigo_analise para o casing canônico da aba Tempo.
Uso: py scripts/extract_xlsm.py [caminho_xlsm]
"""
import sys, csv, datetime as dt
from pathlib import Path
from openpyxl import load_workbook

XLSM = sys.argv[1] if len(sys.argv) > 1 else r"D:\Dropbox\ATGC\Custos\1-Laboratorio\Laboratorio1.xlsm"
OUT = Path(__file__).resolve().parent.parent / "seed"
OUT.mkdir(exist_ok=True)

wb = load_workbook(XLSM, data_only=True)

def rows(sheet, min_row=2):
    return wb[sheet].iter_rows(min_row=min_row, values_only=True)

def s(v):
    return None if v is None else str(v).strip()

def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None

def d(v):
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return dt.datetime.strptime(v.strip(), fmt).date().isoformat()
            except ValueError:
                pass
    return None

def yn(v):
    return s(v) is not None and s(v).lower().startswith("s")

# ---- mapa canônico de codigo_analise (casing da aba Tempo) ----
canon = {}
for r in rows("Tempo"):
    code = s(r[1])
    if code:
        canon.setdefault(code.lower(), code)

def code_norm(v):
    cv = s(v)
    return canon.get(cv.lower(), cv) if cv else None

INSUMO_ALIASES = {
    "QuantiNova® SYBR®Green RT-PCR Kit": "QuantiNova® SYBR®Green RT-PCR Kit (2500 reações)",
    "QuantiNovaŽ SYBRŽGreen RT-PCR Kit": "QuantiNova® SYBR®Green RT-PCR Kit (2500 reações)",
    "dNTP mix 10 mM. Kit c/ 800 uL (4 x 200 uL)": "dNTP mix 10 mM - 1 mL",
}

ANALISE_META = {
    "Eletrof_vir_hem": ("Eletroforese hemolinfa", "Gel para hemolinfa", "Ainda é feito", True),
    "Eletrof_vir_tec": ("Eletroforese tecido", "Gel para tecido", "Ainda é feito", True),
    "Illumina_16S_AC": ("16S alta cobertura", "Sequenciamento focado em microbioma, com alta cobertura", "Nunca feito; pode ser oferecido", True),
    "Illumina_16S_BC": ("16S baixa cobertura", "Sequenciamento focado em microbioma, com baixa cobertura", "Nunca feito; revisar oferta", True),
    "Illumina_DNA_P_AC": ("DNA parasito alta cobertura", "Sequenciamento de DNA de parasito com alta cobertura", "Nunca feito; revisar oferta", True),
    "Illumina_DNA_P_BC": ("DNA parasito baixa cobertura", "Sequenciamento de DNA de parasito com baixa cobertura", "Nunca feito; revisar oferta", True),
    "Illumina_Sh": ("Shotgun", "Sequenciamento shotgun, com qualquer marcador", "Revisar", True),
    "Illumina_Sh_qPCR": ("Shotgun com qPCR", "Shotgun substituindo algumas etapas por qPCR para otimização de tempo e custo", "Ainda não testado", True),
    "qPCR_F": ("qPCR com filtração", "PCR em tempo real com filtração", "Ativo", True),
    "qPCR_SF": ("qPCR sem filtração", "PCR em tempo real sem filtração", "Ativo", True),
    "RTqPCR_RNA_virus_H": ("RT-qPCR vírus hemolinfa", "PCR em tempo real de vírus 1", "Ativo", True),
    "RTqPCR_RNA_virus_T": ("RT-qPCR vírus tecidos", "PCR em tempo real de vírus 2", "Ativo", True),
    "Sanger": ("Sanger", "Sequenciamento Sanger", "Ainda pode ser oferecido", True),
}

def spec_lookup(v):
    sv = s(v)
    return INSUMO_ALIASES.get(sv, sv)

def sqlstr(v):
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(v)
    return "'" + str(v).replace("'", "''") + "'"

def emit(f, table, cols, data):
    if not data:
        return
    f.write(f"\n-- {table} ({len(data)} linhas)\n")
    for row in data:
        vals = ", ".join(sqlstr(x) for x in row)
        f.write(f"insert into {table} ({', '.join(cols)}) values ({vals});\n")

# ---------------- coleta ----------------
analises = sorted(set(canon.values()))
analises_rows = [
    (codigo, *(ANALISE_META.get(codigo) or (codigo, None, "Revisar", True)))
    for codigo in analises
]

etapas = []
for r in rows("Tempo"):
    code = code_norm(r[1])
    if not code:
        continue
    etapas.append((code, s(r[2]), s(r[3]), num(r[4]), num(r[5]),
                   num(r[7]), num(r[8]), yn(r[10]), s(r[11]), s(r[12]), num(r[13])))

equip = []
equip_names = []
for r in rows("Equipamentos"):
    nome = s(r[0])
    if not nome:
        continue
    equip_names.append(nome)
    equip.append((nome, num(r[1]), num(r[2]), d(r[3]), yn(r[5]),
                  num(r[6]), num(r[7]) or 0))

equip_analise = []
for r in rows("Equipamento_Analise"):
    nome = s(r[0]); code = code_norm(r[1])
    if nome and code:
        equip_analise.append((nome, code, num(r[2]) or 0))

tecnicos = []
for r in rows("Tecnicos"):
    nome = s(r[1])
    if nome:
        tecnicos.append((nome, s(r[0]), num(r[2]) or 0, num(r[3]) or 170, num(r[5]) or 0))

overhead = []
for r in rows("Overhead"):
    item = s(r[0])
    if item:
        overhead.append((item, num(r[1]) or 0, num(r[2]) or 100, num(r[3]) or 450))

# MC: chave = coluna B (especificacao). Dedup por especificacao.
insumos = []
seen_spec = set()
for r in rows("MC"):
    spec = s(r[1])
    if not spec or spec in seen_spec:
        continue
    seen_spec.add(spec)
    insumos.append((s(r[0]), spec, num(r[8]), num(r[4]), s(r[6]), num(r[2]), d(r[10])))

insumo_analise = []
for r in rows("MCA"):
    code = code_norm(r[1])
    if not code:
        continue
    insumo_analise.append((code, s(r[2]), s(r[3]), s(r[5]), s(r[7]),
                           s(r[8]), num(r[9]), s(r[11])))

# parametros default (vindos da planilha)
parametros = [
    ("dias_uteis_ano", 222, "dias", "Dias úteis/ano p/ rateio de equipamentos"),
    ("horas_mes_tecnico", 170, "h", "Horas-base mensais por técnico"),
    ("horas_bancada_mes", 450, "h", "Horas de bancada/mês p/ rateio de overhead"),
    ("margem_lucro", 0, "%", "Margem de lucro sobre o custo total"),
    ("impostos", 0, "%", "Impostos sobre a venda"),
    ("taxas", 0, "%", "Taxas administrativas"),
    ("fundo_reserva", 0, "%", "Fundo de reserva"),
    ("fundo_investimento", 0, "%", "Fundo de investimento"),
]

# ---------------- CSVs (para conferência) ----------------
def write_csv(name, header, data):
    with open(OUT / f"{name}.csv", "w", newline="", encoding="utf-8") as fp:
        w = csv.writer(fp); w.writerow(header); w.writerows(data)

write_csv("analises", ["codigo","nome_simplificado","descricao","status","ativo"], analises_rows)
write_csv("etapas", ["codigo_analise","nome_etapa","nome_atividade","execucoes_por_dia",
    "amostras_por_execucao","tempo_maquina_h","tempo_bancada_h","atividade_opcional",
    "tipo_limitacao","dia_inicio","dia_fim_max"], etapas)
write_csv("equipamentos", ["nome","quantidade","custo_unitario","data_aquisicao","possui",
    "vida_util_anos","percentual_manutencao_anual"], equip)
write_csv("equipamento_analise", ["equipamento","codigo_analise","peso_alocacao"], equip_analise)
write_csv("tecnicos", ["nome","processo","valor_mes","horas_mes_base","percentual_dedicado"], tecnicos)
write_csv("overhead", ["item","custo_mensal","percentual_compensada","horas_bancada_mes"], overhead)
write_csv("insumos", ["nome_item","especificacao","custo_total_embalagem","quantidade_embalagem",
    "unidade","custo_unitario","data_aquisicao"], insumos)
write_csv("insumo_analise", ["codigo_analise","nome_etapa","nome_atividade","especificacao_insumo",
    "unidade","grupo_escolha","quantidade_por_amostra","modo_cobranca"], insumo_analise)

# ---------------- seed.sql ----------------
with open(OUT / "seed.sql", "w", encoding="utf-8") as f:
    f.write("-- Seed gerado por scripts/extract_xlsm.py\nbegin;\n")
    emit(f, "parametros", ["chave","valor","unidade","descricao"], parametros)
    emit(f, "analises", ["codigo","nome_simplificado","descricao","status","ativo"], analises_rows)
    emit(f, "etapas", ["codigo_analise","nome_etapa","nome_atividade","execucoes_por_dia",
        "amostras_por_execucao","tempo_maquina_h","tempo_bancada_h","atividade_opcional",
        "tipo_limitacao","dia_inicio","dia_fim_max"], etapas)
    emit(f, "equipamentos", ["nome","quantidade","custo_unitario","data_aquisicao","possui",
        "vida_util_anos","percentual_manutencao_anual"], equip)
    emit(f, "tecnicos", ["nome","processo","valor_mes","horas_mes_base","percentual_dedicado"], tecnicos)
    emit(f, "overhead", ["item","custo_mensal","percentual_compensada","horas_bancada_mes"], overhead)
    emit(f, "insumos", ["nome_item","especificacao","custo_total_embalagem","quantidade_embalagem",
        "unidade","custo_unitario","data_aquisicao"], insumos)
    # equipamento_analise: resolve nome -> id
    f.write("\n-- equipamento_analise\n")
    for nome, code, peso in equip_analise:
        f.write("insert into equipamento_analise (equipamento_id, codigo_analise, peso_alocacao) "
                f"select id, {sqlstr(code)}, {sqlstr(peso)} from equipamentos where nome = {sqlstr(nome)};\n")
    # insumo_analise: resolve especificacao -> insumo_id (left)
    f.write("\n-- insumo_analise\n")
    for code, et, at, spec, un, grp, qpa, modo in insumo_analise:
        f.write("insert into insumo_analise (codigo_analise, nome_etapa, nome_atividade, "
                "especificacao_insumo, unidade, grupo_escolha, quantidade_por_amostra, modo_cobranca, insumo_id) "
                f"select {sqlstr(code)}, {sqlstr(et)}, {sqlstr(at)}, {sqlstr(spec)}, {sqlstr(un)}, "
                f"{sqlstr(grp)}, {sqlstr(qpa)}, {sqlstr(modo)}, "
                f"(select id from insumos where especificacao = {sqlstr(spec_lookup(spec))});\n")
    f.write("\ncommit;\n")

# ---------------- relatório ----------------
specs_mc = seen_spec
specs_mca = {spec_lookup(r[5]) for r in rows("MCA") if s(r[5])}
faltam = sorted(specs_mca - specs_mc)
print(f"analises={len(analises)} etapas={len(etapas)} equipamentos={len(equip)} "
      f"equip_analise={len(equip_analise)} tecnicos={len(tecnicos)} overhead={len(overhead)} "
      f"insumos={len(insumos)} insumo_analise={len(insumo_analise)}")
print(f"insumos referenciados na MCA sem correspondência na MC: {len(faltam)}")
for x in faltam[:40]:
    print("  [FALTA]", x)
